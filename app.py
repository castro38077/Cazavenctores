import pymysql
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
from datetime import date
from openpyxl import Workbook
from io import BytesIO
import os
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cazavectores_secret')

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
app.secret_key = "cazavectores_secret"

# ---------------- CONEXIÓN MYSQL ----------------

def get_connection():
    return pymysql.connect(
        host=os.environ.get('DB_HOST', 'localhost'),
        user=os.environ.get('DB_USER', 'root'),
        password=os.environ.get('DB_PASSWORD', ''),
        database=os.environ.get('DB_NAME', 'cazavectores'),
        cursorclass=pymysql.cursors.DictCursor
    )

# ---------------- LOGIN ----------------

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        connection = get_connection()
        cursor = connection.cursor()

        cursor.execute(
            "SELECT * FROM usuarios WHERE username=%s AND password=%s",
            (username, password),
        )

        user = cursor.fetchone()
        connection.close()

        if user:
            session["logged_in"] = True
            session["username"] = username
            return redirect("/inventario")
        else:
            return render_template("login.html", error="Usuario o contraseña incorrectos")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/")
def home():
    return redirect("/login")

# ---------------- CREAR USUARIOS ----------------
@app.route("/registro", methods=["GET","POST"])
def registro():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]
        clave_admin = request.form["clave_admin"]

        # clave para permitir crear usuarios
        if clave_admin != "admin123":
            return render_template("registro.html", error="Clave de administrador incorrecta")

        connection = get_connection()
        cursor = connection.cursor()

        cursor.execute(
            "INSERT INTO usuarios(username,password) VALUES(%s,%s)",
            (username,password)
        )

        connection.commit()
        connection.close()

        return redirect("/login")

    return render_template("registro.html")

# ---------------- INVENTARIO ----------------

@app.route("/inventario")
def index():

    if not session.get("logged_in"):
        return redirect("/login")

    filter_estado = request.args.get('filter', 'todos')
    page = int(request.args.get('page', 1))
    per_page = 10  # productos por página

    connection = get_connection()
    cursor = connection.cursor()

    # PRODUCTOS AGRUPADOS con filtro
    query = """
        SELECT 
            MIN(id) as id,
            nombre,
            SUM(cantidad) as cantidad,
            precio,
            MIN(fecha_vencimiento) as fecha_vencimiento
        FROM productos
        WHERE cantidad > 0
    """
    params = []

    if filter_estado != 'todos':
        if filter_estado == 'vencido':
            query += " AND DATEDIFF(fecha_vencimiento, CURDATE()) < 0"
        elif filter_estado == 'alerta':
            query += " AND DATEDIFF(fecha_vencimiento, CURDATE()) BETWEEN 0 AND 7"
        elif filter_estado == 'normal':
            query += " AND (DATEDIFF(fecha_vencimiento, CURDATE()) > 7 OR fecha_vencimiento IS NULL)"
        elif filter_estado == 'stock_bajo':
            # Para stock bajo, filtrar después en Python
            pass

    query += " GROUP BY nombre, precio ORDER BY fecha_vencimiento ASC"

    cursor.execute(query, params)
    all_productos = cursor.fetchall()

    # Filtrar stock bajo si necesario
    if filter_estado == 'stock_bajo':
        all_productos = [p for p in all_productos if p.get("cantidad", 0) <= 5]

    # Calcular estados y filtrar según el modo seleccionado (vencido/alerta/normal/stock_bajo)
    hoy = date.today()
    for p in all_productos:
        if p.get("fecha_vencimiento"):
            dias = (p["fecha_vencimiento"] - hoy).days
            p["dias"] = dias
            if dias < 0:
                p["estado"] = "vencido"
            elif dias <= 7:
                p["estado"] = "alerta"
            else:
                p["estado"] = "normal"
        else:
            p["estado"] = "normal"
            p["dias"] = None
        if p["cantidad"] <= 5:
            p["stock_alerta"] = True
        else:
            p["stock_alerta"] = False

    if filter_estado == 'stock_bajo':
        all_productos = [p for p in all_productos if p.get("cantidad", 0) <= 5]
    elif filter_estado == 'normal':
        # Excluir lote con stock bajo (<=5) del listado "normal"
        all_productos = [p for p in all_productos if p.get("cantidad", 0) > 5]

    # Paginación
    total_productos = len(all_productos)
    start = (page - 1) * per_page
    end = start + per_page
    productos = all_productos[start:end]

    # calcular productos en stock bajo (<= 5 unidades) para dashboard
    productos_bajo_stock = sum(1 for p in all_productos if p.get("cantidad", 0) <= 5)

    # DASHBOARD (usar all_productos para métricas totales)

    cursor.execute("SELECT SUM(cantidad * precio) as valor_total FROM productos WHERE cantidad > 0")
    valor_total = cursor.fetchone()["valor_total"] or 0

    cursor.execute("""
    SELECT SUM(cantidad) as por_vencer
    FROM productos
    WHERE cantidad > 0
    AND DATEDIFF(fecha_vencimiento, CURDATE()) BETWEEN 0 AND 7
    """)

    por_vencer = cursor.fetchone()["por_vencer"] or 0

    cursor.execute("""
    SELECT SUM(cantidad) as vencidos
    FROM productos
    WHERE cantidad > 0
    AND DATEDIFF(fecha_vencimiento, CURDATE()) < 0
    """)

    vencidos = cursor.fetchone()["vencidos"] or 0

    cursor.execute("SELECT SUM(cantidad) as total_productos FROM productos WHERE cantidad > 0")
    total_productos_db = cursor.fetchone()["total_productos"] or 0

    # OBTENER TODOS LOS LOTES (filtrar por productos visibles)
    nombres_visibles = [p["nombre"] for p in productos]
    if nombres_visibles:
        placeholders = ','.join(['%s'] * len(nombres_visibles))
        cursor.execute(f"""
            SELECT id, nombre, cantidad, fecha_vencimiento
            FROM productos
            WHERE cantidad > 0 AND nombre IN ({placeholders})
            ORDER BY nombre, fecha_vencimiento
        """, nombres_visibles)
        lotes = cursor.fetchall()
    else:
        lotes = []

    connection.close()

    # Calcular páginas
    total_pages = (total_productos + per_page - 1) // per_page

    return render_template(
        "index.html",
        productos=productos,
        lotes=lotes,
        valor_total=valor_total,
        por_vencer=por_vencer,
        vencidos=vencidos,
        total_productos=total_productos_db,
        productos_bajo_stock=productos_bajo_stock,
        error=request.args.get('error'),
        filter_estado=filter_estado,
        page=page,
        total_pages=total_pages
    )

# ---------------- AGREGAR PRODUCTO ----------------

@app.route("/agregar", methods=["POST"])
def agregar():

    if not session.get("logged_in"):
        return redirect("/login")

    nombre = request.form["nombre"]
    cantidad = int(request.form["cantidad"])
    precio = float(request.form["precio"])
    fecha_vencimiento = request.form["fecha_vencimiento"]

    # Validaciones
    if cantidad <= 0:
        return redirect("/inventario?error=Cantidad debe ser mayor a 0")
    if precio <= 0:
        return redirect("/inventario?error=Precio debe ser mayor a 0")
    if not fecha_vencimiento:
        return redirect("/inventario?error=Fecha de vencimiento es obligatoria")

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        INSERT INTO productos
        (nombre, cantidad, precio, fecha_vencimiento)
        VALUES (%s,%s,%s,%s)
        """,
        (nombre, cantidad, precio, fecha_vencimiento)
    )

    producto_id = cursor.lastrowid

    usuario = session["username"]

    cursor.execute(
        """
        INSERT INTO movimientos (producto_id, tipo, cantidad, usuario)
        VALUES (%s,'entrada',%s,%s)
        """,
        (producto_id, cantidad, usuario)
    )

    cursor.execute(
        """
        INSERT INTO mensajes(usuario, mensaje)
        VALUES(%s,%s)
        """,
        (usuario, f"agregó {cantidad} unidades de {nombre}")
    )

    connection.commit()
    connection.close()
    

    return redirect("/inventario")


# ---------------- VENDER PRODUCTO (FEFO) ----------------

@app.route("/vender", methods=["POST"])
def vender():

    if not session.get("logged_in"):
        return redirect("/login")

    producto_id = request.form["producto_id"]
    cantidad_vender = int(request.form["cantidad"])

    if cantidad_vender <= 0:
        return redirect("/inventario?error=Cantidad a vender debe ser mayor a 0")

    connection = get_connection()
    cursor = connection.cursor()

    usuario = session["username"]

    # Obtener nombre del producto
    cursor.execute(
        "SELECT nombre FROM productos WHERE id=%s",
        (producto_id,)
    )

    producto = cursor.fetchone()
    nombre_producto = producto["nombre"]

    # Calcular stock total disponible para este producto
    cursor.execute(
        "SELECT SUM(cantidad) as stock_total FROM productos WHERE nombre=%s AND cantidad > 0",
        (nombre_producto,)
    )

    stock_total = cursor.fetchone()["stock_total"] or 0

    if cantidad_vender > stock_total:
        connection.close()
        return redirect("/inventario?error=No hay suficiente stock disponible")

    # Buscar lotes ordenados por vencimiento (FEFO)
    cursor.execute(
        """
        SELECT * FROM productos
        WHERE nombre=%s AND cantidad > 0
        ORDER BY fecha_vencimiento ASC
        """,
        (nombre_producto,)
    )

    lotes = cursor.fetchall()

    for lote in lotes:

        if cantidad_vender <= 0:
            break

        stock = lote["cantidad"]
        producto_id_lote = lote["id"]

        if stock <= cantidad_vender:

            cursor.execute(
                "UPDATE productos SET cantidad=0 WHERE id=%s",
                (producto_id_lote,)
            )

            cursor.execute(
                """
                INSERT INTO movimientos (producto_id,tipo,cantidad,usuario)
                VALUES (%s,'salida',%s,%s)
                """,
                (producto_id_lote, stock, usuario)
            )

            cantidad_vender -= stock

        else:

            nuevo_stock = stock - cantidad_vender

            cursor.execute(
                "UPDATE productos SET cantidad=%s WHERE id=%s",
                (nuevo_stock, producto_id_lote)
            )

            cursor.execute(
                """
                INSERT INTO movimientos (producto_id,tipo,cantidad,usuario)
                VALUES (%s,'salida',%s,%s)
                """,
                (producto_id_lote, cantidad_vender, usuario)
            )

            cantidad_vender = 0

    # Registrar mensaje en el chat
    cursor.execute(
        """
        INSERT INTO mensajes(usuario, mensaje)
        VALUES(%s,%s)
        """,
        (usuario, f"{usuario} vendió {nombre_producto} - {request.form['cantidad']} unidades")
    )

    connection.commit()
    connection.close()

    return redirect("/inventario")


# ---------------- HISTORIAL MOVIMIENTOS ----------------

@app.route("/movimientos")
def movimientos():

    if not session.get("logged_in"):
        return redirect("/login")

    tipo = request.args.get("tipo", "todos")
    fecha_desde = request.args.get("desde", "")
    fecha_hasta = request.args.get("hasta", "")
    usuario = request.args.get("usuario", "todos")

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT DISTINCT usuario FROM movimientos ORDER BY usuario")
    usuarios = [row["usuario"] for row in cursor.fetchall()]

    query = """
        SELECT movimientos.*, productos.nombre
        FROM movimientos
        JOIN productos ON movimientos.producto_id = productos.id
        WHERE 1=1
    """
    params = []

    if tipo in ("entrada", "salida"):
        query += " AND movimientos.tipo = %s"
        params.append(tipo)

    if fecha_desde:
        query += " AND movimientos.fecha >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        query += " AND movimientos.fecha <= %s"
        params.append(fecha_hasta)

    if usuario != "todos":
        query += " AND movimientos.usuario = %s"
        params.append(usuario)

    query += " ORDER BY fecha DESC"

    cursor.execute(query, params)
    datos = cursor.fetchall()

    connection.close()

    return render_template(
        "movimientos.html",
        datos=datos,
        tipo=tipo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        usuarios=usuarios,
        usuario=usuario,
    )


@app.route("/movimientos/export")
def export_movimientos():

    if not session.get("logged_in"):
        return redirect("/login")

    tipo = request.args.get("tipo", "todos")
    fecha_desde = request.args.get("desde", "")
    fecha_hasta = request.args.get("hasta", "")
    usuario = request.args.get("usuario", "todos")

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT movimientos.*, productos.nombre
        FROM movimientos
        JOIN productos ON movimientos.producto_id = productos.id
        WHERE 1=1
    """
    params = []

    if tipo in ("entrada", "salida"):
        query += " AND movimientos.tipo = %s"
        params.append(tipo)

    if fecha_desde:
        query += " AND movimientos.fecha >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        query += " AND movimientos.fecha <= %s"
        params.append(fecha_hasta)

    if usuario != "todos":
        query += " AND movimientos.usuario = %s"
        params.append(usuario)

    query += " ORDER BY fecha DESC"

    cursor.execute(query, params)
    datos = cursor.fetchall()

    connection.close()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Headers
    ws['A1'] = 'Producto'
    ws['B1'] = 'Tipo'
    ws['C1'] = 'Cantidad'
    ws['D1'] = 'Usuario'
    ws['E1'] = 'Fecha'

    # Datos
    for i, m in enumerate(datos, start=2):
        ws[f'A{i}'] = m['nombre']
        ws[f'B{i}'] = m['tipo']
        ws[f'C{i}'] = m['cantidad']
        ws[f'D{i}'] = m['usuario']
        ws[f'E{i}'] = str(m['fecha'])

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='movimientos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------- CHAT INTERNO ----------------

@app.route("/chat")
def chat():

    if not session.get("logged_in"):
        return redirect("/login")

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM mensajes ORDER BY fecha DESC")
    mensajes = cursor.fetchall()

    connection.close()

    return render_template("chat.html", mensajes=mensajes)


@app.route("/enviar", methods=["POST"])
def enviar():

    if not session.get("logged_in"):
        return redirect("/login")

    usuario = session["username"]

    if request.is_json:
        data = request.get_json()
        mensaje = data.get("mensaje", "").strip()
    else:
        mensaje = request.form.get("mensaje", "").strip()

    if not mensaje:
        if request.is_json:
            return jsonify({"success": False, "error": "Mensaje vacío"}), 400
        return redirect("/chat")

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        "INSERT INTO mensajes(usuario, mensaje) VALUES(%s,%s)",
        (usuario, mensaje),
    )

    connection.commit()
    connection.close()

    if request.is_json:
        return jsonify({"success": True})

    return redirect("/chat")


# ---------------- EJECUTAR APP ----------------

if __name__ == "__main__":
    app.run(debug=True)

